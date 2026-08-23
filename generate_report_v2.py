# -*- coding: utf-8 -*-
"""2nd iteration report: adds Pros, Cons, Overall Comments, and Rank columns."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = json.load(open("results_data.json", encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "G2 Review Results v2"

headers = ["Rank", "Candidate Folder", "Question ID", "Functional\nCorrectness\n(45)",
           "Problem\nInterpretation\n(10)", "Algorithm &\nData Structures\n(15)",
           "Testing &\nEdge Cases\n(15)", "Code\nQuality\n(15)", "Total\nScore\n(100)",
           "Grade", "AI Used", "Pros", "Cons", "Review Flags", "Overall Comments"]

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap_center
    cell.border = border

ws.freeze_panes = "A2"

def score_color(total):
    if total >= 90:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif total >= 70:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif total >= 50:
        return PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    else:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def grade(total):
    if total >= 90:
        return "A"
    elif total >= 80:
        return "B"
    elif total >= 70:
        return "C"
    elif total >= 50:
        return "D"
    else:
        return "F"

def overall_comment(rec):
    total = rec["total_score"]
    flags = rec.get("review_flags", [])
    critical = [f for f in flags if "CRITICAL" in f.upper()]
    if total >= 95:
        base = "Excellent, near-flawless submission. Strong hire signal for this challenge."
    elif total >= 85:
        base = "Strong submission with correct logic and only minor polish gaps."
    elif total >= 70:
        base = "Solid attempt with correct core logic but noticeable gaps in testing/quality."
    elif total >= 50:
        base = "Partial solution with a significant functional or logic bug; needs follow-up."
    else:
        base = "Weak/failing submission with critical correctness issues."
    if critical:
        base += " Critical concern(s): " + "; ".join(critical) + "."
    if rec.get("ai_used") == "Yes":
        base += " Possible AI assistance signals noted (informational only, not penalized)."
    return base

# Sort by total_score descending for ranking
ranked = sorted(DATA, key=lambda r: r["total_score"], reverse=True)

row_idx = 2
for i, rec in enumerate(ranked, 1):
    cs = rec["criterion_scores"]
    total = rec["total_score"]
    values = [
        i,
        rec["folder"],
        rec["question_id"],
        cs["functional_correctness"],
        cs["problem_interpretation"],
        cs["algorithm_data_structures"],
        cs["testing_edge_cases"],
        cs["code_quality"],
        total,
        grade(total),
        rec.get("ai_used", "No"),
        "; ".join(rec.get("strengths", [])) or "None noted",
        "; ".join(rec.get("improvements", [])) or "None noted",
        "; ".join(rec.get("review_flags", [])) or "None",
        overall_comment(rec),
    ]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.border = border
        if c in (1, 3, 4, 5, 6, 7, 8, 9, 10, 11):
            cell.alignment = wrap_center
        else:
            cell.alignment = wrap_left
        if c in (1, 9, 10):
            cell.fill = score_color(total)
            cell.font = Font(bold=True)
    row_idx += 1

widths = [6, 20, 12, 12, 12, 14, 12, 11, 10, 8, 9, 42, 42, 32, 48]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 45
for r in range(2, row_idx):
    ws.row_dimensions[r].height = 70

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "G2 Coding Challenge Review Summary (2nd Iteration)"
ws2["A1"].font = Font(bold=True, size=14)
totals = [rec["total_score"] for rec in DATA]
ws2["A3"] = "Total Submissions Reviewed"
ws2["B3"] = len(DATA)
ws2["A4"] = "Average Score"
ws2["B4"] = round(sum(totals) / len(totals), 2)
ws2["A5"] = "Highest Score"
ws2["B5"] = max(totals)
ws2["A6"] = "Lowest Score"
ws2["B6"] = min(totals)
ws2["A8"] = "Grade Distribution"
ws2["A8"].font = Font(bold=True)
grades = {}
for t in totals:
    g = grade(t)
    grades[g] = grades.get(g, 0) + 1
r = 9
for g in ["A", "B", "C", "D", "F"]:
    ws2.cell(row=r, column=1, value=g)
    ws2.cell(row=r, column=2, value=grades.get(g, 0))
    r += 1

r += 1
ws2.cell(row=r, column=1, value="Top 5 Candidates (by Total Score)").font = Font(bold=True)
r += 1
for i, rec in enumerate(ranked[:5], 1):
    ws2.cell(row=r, column=1, value=f"{i}. {rec['folder']} ({rec['question_id']})")
    ws2.cell(row=r, column=2, value=rec["total_score"])
    r += 1

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 12

wb.save("G2_Coding_Challenge_Review_Results_v2.xlsx")
print("Saved:", row_idx - 2, "rows (2nd iteration)")
